"""Module to generate sample Excel test spreadsheet for NetBox Device Importer."""
import random
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_template_excel(file_path: str = "netbox_test_devices.xlsx") -> str:
    """Creates a 2-sheet formatted Excel workbook with configuration and test devices.
    
    Sheet 1 ('Config'): Contains NetBox API settings and environment variables.
    Sheet 2 ('Devices'): Contains 10 sample device names and IP addresses (between 192.168.70.1-200).
    """
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Sheet 1: Config
    # ----------------------------------------------------
    ws_config = wb.active
    ws_config.title = "Config"
    
    # Styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    ws_config["A1"] = "NetBox Environment Configuration"
    ws_config["A1"].font = title_font
    ws_config.merge_cells("A1:C1")

    headers_config = ["Setting Name", "Value", "Description / Notes"]
    for col_num, header in enumerate(headers_config, 1):
        cell = ws_config.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    config_rows = [
        ("CONFIG_VERSION", "1.2.0", "Spreadsheet template version compatibility"),
        ("NETBOX_HOST", "192.168.1.172", "NetBox / Malcolm host IP or domain (defaults to localhost if blank)"),
        ("NETBOX_PATH", "/netbox", "URL path prefix for Malcolm NetBox (default /netbox)"),
        ("NETBOX_PORT", "443", "Port number (443 for HTTPS, 80/8000 for HTTP)"),
        ("NETBOX_SCHEME", "https", "http or https"),
        ("NETBOX_TOKEN", "ee9f0df1ea49b6bbfda7462cb21a446d777b2f4d", "NetBox REST API Token"),
        ("SITE_NAME", "Malcolm Site", "NetBox Site name (created if missing)"),
        ("DEVICE_ROLE", "Generic Device", "NetBox Device Role (created if missing)"),
        ("MANUFACTURER", "Unknown", "NetBox Manufacturer (created if missing)"),
        ("DEVICE_TYPE", "Unknown Model", "NetBox Device Type (created if missing)"),
        ("INTERFACE_NAME", "eth0", "Default interface name attached to each device"),
        ("DEFAULT_SUBNET_MASK", "/24", "Fallback subnet mask if not explicitly specified in Devices sheet"),
        ("VERIFY_SSL", "False", "True or False (set False for self-signed certificates)"),
    ]

    for row_num, row_data in enumerate(config_rows, start=4):
        for col_num, val in enumerate(row_data, start=1):
            cell = ws_config.cell(row=row_num, column=col_num, value=val)
            cell.border = thin_border
            if col_num == 1:
                cell.font = Font(bold=True)

    # ----------------------------------------------------
    # Sheet 2: Devices
    # ----------------------------------------------------
    ws_devices = wb.create_sheet(title="Devices")
    
    ws_devices["A1"] = "Target Device Import List"
    ws_devices["A1"].font = title_font
    ws_devices.merge_cells("A1:B1")

    headers_devices = ["Device Name", "IP Address", "Overwrite"]
    for col_num, header in enumerate(headers_devices, 1):
        cell = ws_devices.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Generate 10 sample device names and random IPs between 192.168.70.1 and 192.168.70.200
    random.seed(42)  # Deterministic seed for reproducible test file
    used_octets = set()
    sample_devices = []
    
    # 10 device entries
    device_names = [
        "srv-sensor-01",
        "srv-sensor-02",
        "fw-edge-01",
        "sw-core-01",
        "rt-gateway-01",
        "mon-node-01",
        "mon-node-02",
        "plc-ctrl-01",
        "hmi-term-01",
        "srv-custom-01"
    ]

    for idx, name in enumerate(device_names):
        while True:
            octet = random.randint(1, 200)
            if octet not in used_octets:
                used_octets.add(octet)
                break
        
        # Test case: 9th entry has explicit custom subnet mask /16
        if idx == 8:
            ip_str = f"192.168.70.{octet}/16"
        else:
            ip_str = f"192.168.70.{octet}"
        
        # Test case: 10th entry has Overwrite set to True to demonstrate overwrite functionality
        overwrite_val = "TRUE" if idx == 9 else "FALSE"
        
        sample_devices.append((name, ip_str, overwrite_val))

    for row_num, (dev_name, ip_addr, ow_val) in enumerate(sample_devices, start=4):
        c1 = ws_devices.cell(row=row_num, column=1, value=dev_name)
        c2 = ws_devices.cell(row=row_num, column=2, value=ip_addr)
        c3 = ws_devices.cell(row=row_num, column=3, value=ow_val)
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border

    # ----------------------------------------------------
    # Sheet 3: Prefixes (IPAM Prefixes)
    # ----------------------------------------------------
    ws_prefixes = wb.create_sheet(title="Prefixes")
    
    ws_prefixes["A1"] = "IPAM Prefix Subnet List"
    ws_prefixes["A1"].font = title_font
    ws_prefixes.merge_cells("A1:D1")

    headers_prefixes = ["Prefix", "Description", "Site", "Status"]
    for col_num, header in enumerate(headers_prefixes, 1):
        cell = ws_prefixes.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_prefixes = [
        ("192.168.70.0/24", "Lab Test Devices Subnet", "Malcolm Site", "active"),
        ("192.168.1.0/16", "Corporate Infrastructure Subnet", "Malcolm Site", "active"),
    ]

    for row_num, (prefix_cidr, desc, site, status) in enumerate(sample_prefixes, start=4):
        c1 = ws_prefixes.cell(row=row_num, column=1, value=prefix_cidr)
        c2 = ws_prefixes.cell(row=row_num, column=2, value=desc)
        c3 = ws_prefixes.cell(row=row_num, column=3, value=site)
        c4 = ws_prefixes.cell(row=row_num, column=4, value=status)
        c1.border = thin_border
        c2.border = thin_border
        c3.border = thin_border
        c4.border = thin_border

    # Adjust column widths for visual appeal
    for ws in [ws_config, ws_devices, ws_prefixes]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    target_path = Path(file_path).resolve()
    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    return str(target_path)


if __name__ == "__main__":
    out = create_template_excel()
    print(f"Template created at: {out}")
