"""Pre-flight Validator for NetBox Excel Device Importer."""
import ipaddress
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Any, Optional
import openpyxl


@dataclass
class ValidationResult:
    config: Dict[str, str] = field(default_factory=dict)
    devices: List[Dict[str, str]] = field(default_factory=list)
    prefixes: List[Dict[str, str]] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def is_valid(self) -> bool:
        return len(self.errors) == 0


def _clean_str(val: Any) -> str:
    """Sanitizes cell input into a stripped string."""
    if val is None:
        return ""
    return str(val).strip()


def validate_excel_file(file_path: str) -> ValidationResult:
    """Parses and validates an Excel file for NetBox device import.
    
    Performs pre-flight checks:
    - Validates presence of Excel file and required sheets.
    - Extracts Config settings with safe defaults (NETBOX_HOST defaults to localhost if empty).
    - Checks device rows for:
        * Missing device names or IP addresses
        * Duplicate device names or IP addresses within the sheet
        * Malformed IP syntax (via ipaddress module)
    """
    result = ValidationResult()
    path = Path(file_path)

    if not path.exists():
        result.errors.append(f"Excel file not found at path: '{file_path}'")
        return result

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        result.errors.append(f"Failed to open Excel file '{file_path}': {str(e)}")
        return result

    sheet_names = wb.sheetnames
    if not sheet_names:
        result.errors.append("Excel workbook is empty and contains no sheets.")
        return result

    # ----------------------------------------------------
    # 1. Parse Config Sheet
    # ----------------------------------------------------
    config_sheet_name = None
    for name in sheet_names:
        if name.strip().lower() in ("config", "configuration", "settings", "env", "environment"):
            config_sheet_name = name
            break
    if not config_sheet_name:
        config_sheet_name = sheet_names[0]  # Fallback to first sheet
        result.warnings.append(f"Config sheet not explicitly named 'Config'. Reading from '{config_sheet_name}'.")

    ws_config = wb[config_sheet_name]
    raw_config: Dict[str, str] = {}

    for row in ws_config.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue
        key_raw = _clean_str(row[0]).upper()
        val_raw = _clean_str(row[1])
        if key_raw:
            raw_config[key_raw] = val_raw

    # Apply defaults
    config_version = raw_config.get("CONFIG_VERSION") or "1.0.0"
    host = raw_config.get("NETBOX_HOST") or "localhost"
    path_prefix = raw_config.get("NETBOX_PATH") or "/netbox"
    port = raw_config.get("NETBOX_PORT") or "443"
    scheme = raw_config.get("NETBOX_SCHEME") or "https"
    token = raw_config.get("NETBOX_TOKEN") or ""
    site = raw_config.get("SITE_NAME") or "Malcolm Site"
    role = raw_config.get("DEVICE_ROLE") or "Generic Device"
    manufacturer = raw_config.get("MANUFACTURER") or "Unknown"
    device_type = raw_config.get("DEVICE_TYPE") or "Unknown Model"
    interface_name = raw_config.get("INTERFACE_NAME") or "eth0"
    default_mask = raw_config.get("DEFAULT_SUBNET_MASK") or "/24"
    if not default_mask.startswith("/"):
        default_mask = f"/{default_mask}"
    
    verify_ssl_str = raw_config.get("VERIFY_SSL") or "False"
    verify_ssl = verify_ssl_str.lower() in ("true", "1", "yes", "t")

    # Build constructed URL if NETBOX_URL is not explicitly set
    explicit_url = raw_config.get("NETBOX_URL")
    if explicit_url:
        netbox_base_url = explicit_url.rstrip("/")
    else:
        clean_path = path_prefix.strip("/")
        path_segment = f"/{clean_path}" if clean_path else ""
        if (scheme == "https" and port == "443") or (scheme == "http" and port == "80"):
            netbox_base_url = f"{scheme}://{host}{path_segment}"
        else:
            netbox_base_url = f"{scheme}://{host}:{port}{path_segment}"

    result.config = {
        "CONFIG_VERSION": config_version,
        "NETBOX_HOST": host,
        "NETBOX_URL": netbox_base_url,
        "NETBOX_TOKEN": token,
        "SITE_NAME": site,
        "DEVICE_ROLE": role,
        "MANUFACTURER": manufacturer,
        "DEVICE_TYPE": device_type,
        "INTERFACE_NAME": interface_name,
        "DEFAULT_SUBNET_MASK": default_mask,
        "VERIFY_SSL": str(verify_ssl),
    }

    if not token:
        result.warnings.append("NETBOX_TOKEN is empty in Config sheet. Ensure token is provided via environment or CLI.")

    # ----------------------------------------------------
    # 2. Parse Devices Sheet
    # ----------------------------------------------------
    devices_sheet_name = None
    for name in sheet_names:
        if name.strip().lower() in ("devices", "device_list", "hosts", "inventory"):
            devices_sheet_name = name
            break
    if not devices_sheet_name:
        if len(sheet_names) > 1:
            devices_sheet_name = sheet_names[1]
        else:
            devices_sheet_name = sheet_names[0]
            result.warnings.append("Single sheet detected; attempting to read devices from second table.")

    ws_devices = wb[devices_sheet_name]
    rows = list(ws_devices.iter_rows(values_only=True))
    if not rows:
        result.errors.append(f"Devices sheet '{devices_sheet_name}' contains no data.")
        return result

    # Locate Header Row
    name_col_idx: Optional[int] = None
    ip_col_idx: Optional[int] = None
    overwrite_col_idx: Optional[int] = None
    header_row_idx: Optional[int] = None

    for r_idx, row in enumerate(rows):
        if not row:
            continue
        row_str = [_clean_str(c).lower() for c in row]
        for c_idx, val in enumerate(row_str):
            if "name" in val or "device" in val:
                if name_col_idx is None:
                    name_col_idx = c_idx
            if "ip" in val or "address" in val:
                if ip_col_idx is None:
                    ip_col_idx = c_idx
            if "overwrite" in val or "force" in val:
                if overwrite_col_idx is None:
                    overwrite_col_idx = c_idx
        if name_col_idx is not None and ip_col_idx is not None:
            header_row_idx = r_idx
            break

    # Fallback to column 0 (Name), column 1 (IP), column 2 (Overwrite) if header detection missed
    if name_col_idx is None or ip_col_idx is None:
        name_col_idx = 0
        ip_col_idx = 1
        overwrite_col_idx = 2
        header_row_idx = 0
        result.warnings.append("Could not automatically locate header row. Assuming Column 1 = Device Name, Column 2 = IP Address, Column 3 = Overwrite.")

    seen_names: Dict[str, int] = {}
    seen_ips: Dict[str, int] = {}

    for r_idx in range(header_row_idx + 1, len(rows)):
        row = rows[r_idx]
        if not row or all(c is None or _clean_str(c) == "" for c in row):
            continue

        excel_row_num = r_idx + 1
        dev_name = _clean_str(row[name_col_idx]) if len(row) > name_col_idx else ""
        raw_ip = _clean_str(row[ip_col_idx]) if len(row) > ip_col_idx else ""
        
        raw_overwrite = ""
        if overwrite_col_idx is not None and len(row) > overwrite_col_idx:
            raw_overwrite = _clean_str(row[overwrite_col_idx]).lower()
        
        overwrite = raw_overwrite in ("true", "1", "yes", "y", "t", "force", "overwrite")

        # Validation Checks
        if not dev_name:
            result.errors.append(f"Row {excel_row_num}: Missing device name.")
        
        if not raw_ip:
            result.errors.append(f"Row {excel_row_num} ('{dev_name or 'Unknown'}'): Missing IP address.")
            continue

        # Check duplicate names
        dev_name_lower = dev_name.lower()
        if dev_name_lower in seen_names:
            prev_row = seen_names[dev_name_lower]
            result.errors.append(f"Row {excel_row_num}: Duplicate device name '{dev_name}' (first seen on Row {prev_row}).")
        else:
            seen_names[dev_name_lower] = excel_row_num

        # Format CIDR Subnet Mask
        if "/" in raw_ip:
            cidr_ip = raw_ip
        else:
            cidr_ip = f"{raw_ip}{default_mask}"

        # Validate IP Syntax using ipaddress module
        try:
            iface = ipaddress.ip_interface(cidr_ip)
            if iface.version != 4:
                result.errors.append(f"Row {excel_row_num} ('{dev_name}'): Only IPv4 addresses are supported ({raw_ip}).")
            else:
                ip_only = str(iface.ip)
                if ip_only in seen_ips:
                    prev_row = seen_ips[ip_only]
                    result.errors.append(f"Row {excel_row_num} ('{dev_name}'): Duplicate IP address '{ip_only}' (first seen on Row {prev_row}).")
                else:
                    seen_ips[ip_only] = excel_row_num

                result.devices.append({
                    "name": dev_name,
                    "raw_ip": raw_ip,
                    "cidr_ip": str(iface),
                    "ip_address": str(iface.ip),
                    "prefix_len": str(iface.network.prefixlen),
                    "overwrite": overwrite,
                    "excel_row": str(excel_row_num),
                })
        except ValueError as ve:
            result.errors.append(f"Row {excel_row_num} ('{dev_name}'): Invalid IP address format '{raw_ip}' ({str(ve)}).")

    # ----------------------------------------------------
    # 3. Parse Prefixes Sheet (If Present)
    # ----------------------------------------------------
    prefix_sheet_name = None
    for name in sheet_names:
        if name.strip().lower() in ("prefixes", "prefix", "ipam", "subnets", "ipam_prefixes"):
            prefix_sheet_name = name
            break

    if prefix_sheet_name:
        ws_pref = wb[prefix_sheet_name]
        p_rows = list(ws_pref.iter_rows(values_only=True))
        if p_rows:
            p_prefix_col: Optional[int] = None
            p_desc_col: Optional[int] = None
            p_site_col: Optional[int] = None
            p_status_col: Optional[int] = None
            p_header_row: Optional[int] = None

            for r_idx, row in enumerate(p_rows):
                if not row:
                    continue
                r_str = [_clean_str(c).lower() for c in row]
                # Check if this row is a header row (must contain 'prefix' in a cell and not be a merged title row)
                for c_idx, val in enumerate(r_str):
                    if val in ("prefix", "subnet", "subnet prefix", "network"):
                        p_prefix_col = c_idx
                    if "desc" in val or "note" in val:
                        p_desc_col = c_idx
                    if "site" in val:
                        p_site_col = c_idx
                    if "status" in val:
                        p_status_col = c_idx
                if p_prefix_col is not None and (p_desc_col is not None or p_site_col is not None or p_status_col is not None):
                    p_header_row = r_idx
                    break

            if p_prefix_col is None:
                p_prefix_col = 0
                p_desc_col = 1
                p_site_col = 2
                p_status_col = 3
                p_header_row = 0

            seen_prefixes: Dict[str, int] = {}
            for r_idx in range(p_header_row + 1, len(p_rows)):
                row = p_rows[r_idx]
                if not row or all(c is None or _clean_str(c) == "" for c in row):
                    continue

                excel_row_num = r_idx + 1
                raw_pref = _clean_str(row[p_prefix_col]) if len(row) > p_prefix_col else ""
                desc = _clean_str(row[p_desc_col]) if p_desc_col is not None and len(row) > p_desc_col else ""
                pref_site = _clean_str(row[p_site_col]) if p_site_col is not None and len(row) > p_site_col else ""
                pref_status = _clean_str(row[p_status_col]) if p_status_col is not None and len(row) > p_status_col else "active"

                if not raw_pref:
                    result.errors.append(f"Prefixes Sheet Row {excel_row_num}: Missing prefix value.")
                    continue

                try:
                    net_obj = ipaddress.ip_network(raw_pref, strict=False)
                    norm_pref = str(net_obj)
                    if norm_pref in seen_prefixes:
                        prev_r = seen_prefixes[norm_pref]
                        result.errors.append(f"Prefixes Sheet Row {excel_row_num}: Duplicate prefix '{norm_pref}' (first seen on Row {prev_r}).")
                    else:
                        seen_prefixes[norm_pref] = excel_row_num

                    result.prefixes.append({
                        "prefix": norm_pref,
                        "description": desc,
                        "site": pref_site or site,  # Fallback to Config site if blank
                        "status": pref_status or "active",
                        "excel_row": str(excel_row_num),
                    })
                except ValueError as ve:
                    result.errors.append(f"Prefixes Sheet Row {excel_row_num}: Invalid subnet prefix format '{raw_pref}' ({str(ve)}).")

    return result
